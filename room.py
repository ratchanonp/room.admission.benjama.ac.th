import pandas as pd
import numpy as np
import itertools
import openpyxl
import logging
import sys
from typing import Dict, List, Tuple, Optional
from pathlib import Path
from fpdf import FPDF


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('exam_room_assignment.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Constants
EXCEL_INPUT_FILE = "ApplicantNamelist.xlsx"
SHEET_NAME = "Merge"
EXAM_ID_PREFIX = {
    "m1-special-epsmtp": "11",
    "m1-special-smte": "12",
    "m4-special-smte": "41",
    "m4-special-hsip": "42",
    "m4-special-ep-scimath": "43",
    "m4-special-ep-artmath": "44",
    "m4-special-lang-cn": "46",
    "m4-special-lang-jp": "45",
    "m4-special-lang-fr": "47",
}

PROGRAM_THAI_NAME = {
    "m1-special-epsmtp": "โครงการจัดการเรียนการสอนตามหลักสูตรกระทรวงศึกษาธิการเป็นภาษาอังกฤษ และโครงการส่งเสริมความเป็นเลิศด้านวิทยาศาสตร์ คณิตศาสตร์ และเทคโนโลยี",
    "m1-special-smte": "โครงการห้องเรียนพิเศษวิทยาศาสตร์ คณิตศาสตร์ เทคโนโลยีและสิ่งแวดล้อม ระดับมัธยมศึกษาตอนต้น",
    "m4-special-smte": "โครงการห้องเรียนพิเศษวิทยาศาสตร์ คณิตศาสตร์ เทคโนโลยีและสิ่งแวดล้อม",
    "m4-special-hsip": "โครงการห้องเรียนพิเศษวิทยาศาสตร์สุขภาพ",
    "m4-special-ep-scimath": "โครงการจัดการเรียนการสอนตามหลักสูตรกระทรวงศึกษาธิการเป็นภาษาอังกฤษ แผนการเรียนวิทย์-คณิต",
    "m4-special-ep-artmath": "	โครงการจัดการเรียนการสอนตามหลักสูตรกระทรวงศึกษาธิการเป็นภาษาอังกฤษ แผนการเรียนศิลป์-คำนวณ",
    "m4-special-lang-cn": "โครงการห้องเรียนพิเศษภาษาต่างประเทศภาษาที่ 2 แผนการเรียนศิลป์ - ภาษาจีน",
    "m4-special-lang-jp": "โครงการห้องเรียนพิเศษภาษาต่างประเทศภาษาที่ 2 แผนการเรียนศิลป์ - ภาษาญี่ปุ่น",
    "m4-special-lang-fr": "โครงการห้องเรียนพิเศษภาษาต่างประเทศภาษาที่ 2 แผนการเรียนศิลป์ - ภาษาฝรั่งเศส",
}

TITLE_MAPPING = {
    "นาย": "นาย",
    "นาง": "นาง",
    "นางสาว": "นางสาว",
    "เด็กชาย": "ด.ช.",
    "เด็กหญิง": "ด.ญ."
}

SCHOOL_PREFIXES = [
    "โรงเรียน", "รร.", "ร.ร.", "ร.ร", 
    "โรงเรีย", "โรวเรียน", "เรียน"
]

EXAM_ROOMS = {
    "m1-special-epsmtp": [
    ("721", 30),
    ("722", 30),
    ("723", 30),
    ("724", 30),
    ("725", 30),
    ("726", 30),
    ("727", 30),
    ("728", 30),
    ("742", 30),
    ("743", 30),
    ("744", 30),
    ("745", 30),
    ("746", 30),
    ("747", 30),
    ("748", 30),
    ("ประดู่แดง 1", 30),
    ("ประดู่แดง 2", 30),
    ("ประดู่แดง 3", 30),
    ("622", 30),
    ("632", 30),
    ("633", 30),
    ("634", 30),
    ("635", 30),
    ("641", 30),
    ("642", 30),
    ("643", 30),
    ("644", 30),
    ("645", 30),
    ("636", 30)]    ,
    "m1-special-smte": [
    ("321", 30),
    ("322", 30),
    ("323", 30),
    ("331", 30),
    ("332", 30),
    ("333", 30)],
    "m4-special-smte": [("721", 30), ("722", 30), ("723", 30), ("724", 30), ("725", 30), 
                        ("726", 30), ("727", 30), ("728", 30)],
    "m4-special-hsip": [("742", 30), ("743", 30), ("744", 30), ("745", 30), ("746", 30), ("747", 30), ("748", 30), ("622", 30)],
    "m4-special-ep-scimath": [("ประดู่แดง 1", 30), ("ประดู่แดง 2", 30), ("ประดู่แดง 3", 30), ("632", 30), ("633", 30)],
    "m4-special-ep-artmath": [("634", 32), ("635", 32), ("641", 33)],
    "m4-special-lang-cn": [("642", 32), ("643", 32), ("644", 31)],
    "m4-special-lang-jp": [("645", 32), ("เกษตร", 31)],
    "m4-special-lang-fr": [("สะเต็ม", 30), ("ชย.", 17)],
}

REQUIRED_COLUMNS = [
    "applicant.thaiID", "applicant.title", "applicant.firstName",
    "applicant.lastName", "education.currentSchool", "programID", "status"
]

COLUMN_RENAME_MAPPING = {
    "applicant.thaiID": "thaiID",
    "applicant.title": "title",
    "applicant.firstName": "firstname",
    "applicant.lastName": "lastname",
    "education.currentSchool": "school",
    "programID": "program",
    "status": "status"
}

FINAL_COLUMN_RENAME = {
    "exam_room": "ห้องสอบ",
    "exam_no": "เลขที่นั่งสอบ",
    "exam_id": "เลขประจำตัวสอบ",
    "fullname": "ชื่อ - นามสกุล",
    "school": "โรงเรียน"
}

def validate_input_file(file_path: str) -> bool:
    """Validate input file existence and format"""
    try:
        if not Path(file_path).exists():
            logger.error(f"Input file {file_path} does not exist")
            return False
        
        df = pd.read_excel(file_path, sheet_name=SHEET_NAME, nrows=1)
        missing_cols = set(REQUIRED_COLUMNS) - set(df.columns)
        if missing_cols:
            logger.error(f"Missing required columns: {missing_cols}")
            return False
        return True
    except Exception as e:
        logger.error(f"Error validating input file: {str(e)}")
        return False

def load_and_prepare_data(file_path: str, sheet_name: str) -> Optional[pd.DataFrame]:
    """Load data from Excel file and prepare initial dataframe"""
    try:
        if not validate_input_file(file_path):
            return None
            
        logger.info(f"Loading data from {file_path}")
        data = pd.read_excel(file_path, sheet_name=sheet_name)
        data = data[REQUIRED_COLUMNS]
        data = data.rename(columns=COLUMN_RENAME_MAPPING)
        data["thaiID"] = data["thaiID"].astype(str)
        qualified_data = data[data["status"] == "ผ่านคุณสมบัติ"]
        
        logger.info(f"Loaded {len(qualified_data)} qualified applicants")
        return qualified_data
    except Exception as e:
        logger.error(f"Error loading data: {str(e)}")
        return None

def clean_school_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean school names by removing prefixes and extra whitespace
    """
    df = df.copy()
    pattern = r"^(" + "|".join(SCHOOL_PREFIXES) + ")"
    df["school"] = df["school"].str.replace(pattern, "", regex=True)
    df["school"] = df["school"].str.strip()
    return df

def format_student_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format student names with proper titles and create fullname column
    """
    df = df.copy()
    df["title"] = df["title"].map(TITLE_MAPPING)
    df["firstname"] = df["firstname"].str.strip()
    df["lastname"] = df["lastname"].str.strip()
    
    df["fullname"] = (df["title"] + df["firstname"] + " " + df["lastname"])
    df["fullname"] = df["fullname"].str.strip().str.replace("  ", " ")
    return df

def assign_exam_details(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """Assign exam IDs, room numbers, and seat numbers"""
    try:
        df = df.copy()
        df = df.sort_values(by=["firstname", "lastname"])
        
        # Initialize columns
        df["running_number"] = range(1, len(df) + 1)
        df["exam_id"] = df["program"].map(EXAM_ID_PREFIX) + df["running_number"].astype(str).str.zfill(3)
        df["exam_room"] = ""
        df["exam_no"] = 0
        
        # Process each program separately
        for program in df["program"].unique():
            try:
                program_mask = df["program"] == program
                num_students = program_mask.sum()
                
                if program not in EXAM_ROOMS:
                    logger.error(f"No room configuration found for program: {program}")
                    continue
                
                # Get program rooms and their capacities
                program_rooms = EXAM_ROOMS[program]
                total_capacity = sum(capacity for _, capacity in program_rooms)
                
                if num_students > total_capacity:
                    logger.warning(f"Program {program} has more students ({num_students}) than capacity ({total_capacity})")
                
                # Assign room numbers and seat numbers
                current_student = 0
                room_assignments = []
                exam_numbers = []
                
                for room_name, room_capacity in program_rooms:
                    if current_student >= num_students:
                        break
                    students_in_room = min(room_capacity, num_students - current_student)
                    room_assignments.extend([room_name] * students_in_room)
                    exam_numbers.extend(range(1, students_in_room + 1))
                    current_student += students_in_room
                    logger.debug(f"Assigned {students_in_room} students to room {room_name}")
                
                df.loc[program_mask, "exam_room"] = room_assignments
                df.loc[program_mask, "exam_no"] = exam_numbers
                logger.info(f"Successfully assigned rooms for program {program} ({num_students} students)")
                
            except Exception as e:
                logger.error(f"Error processing program {program}: {str(e)}")
                continue
        
        return df
    except Exception as e:
        logger.error(f"Error in exam detail assignment: {str(e)}")
        return None

def prepare_final_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare the final dataframe for export
    """
    final_columns = ["exam_room", "exam_no", "exam_id", "fullname", "school"]
    df = df[final_columns]
    return df.rename(columns=FINAL_COLUMN_RENAME)

def save_to_excel(df: pd.DataFrame, program: str, output_file: str = "exam_room_assignments.xlsx") -> bool:
    """Save DataFrame to Excel with error handling, each program to a different sheet"""
    try:
        # Check if file exists to determine whether to create new or append
        if Path(output_file).exists():
            mode = 'a'
            if_sheet_exists = 'replace'
        else:
            mode = 'w'
            if_sheet_exists = None
        
        with pd.ExcelWriter(output_file, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
            # Save to sheet named after the program
            df.to_excel(writer, sheet_name=program, index=False)
            
            # Get the worksheet to adjust column widths
            worksheet = writer.sheets[program]
            
            # Set column widths
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                ) + 2
                # Limit maximum column width
                max_length = min(max_length, 50)
                # Convert character units to Excel units (approximate)
                excel_width = max_length * 1.2
                worksheet.column_dimensions[chr(65 + idx)].width = excel_width

        logger.info(f"Successfully saved program {program} to sheet in {output_file}")
        return True
    except Exception as e:
        logger.error(f"Error saving program {program} to Excel {output_file}: {str(e)}")
        return False

def create_pdf_file(df: pd.DataFrame, program: str, output_file: str = "exam_room_assignments.pdf") -> bool:
    """Create a PDF file with exam room assignments for a specific program.
    
    Args:
        df: DataFrame containing exam room assignments
        program: Program ID string
        output_file: Output PDF file path
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Initialize PDF
        pdf = FPDF(orientation='portrait', format='A4')  # Portrait orientation
        pdf.add_font('thsarabun', '', 'fonts/THSarabun.ttf')
        pdf.add_font('thsarabun', 'B', 'fonts/THSarabun Bold.ttf')
        
        # Get rooms in the order defined in EXAM_ROOMS
        if program not in EXAM_ROOMS:
            logger.error(f"No room configuration found for program: {program}")
            return False
            
        # Get all rooms for this program in their original order
        program_room_list = [room_name for room_name, _ in EXAM_ROOMS[program]]
        
        # Get rooms that actually have students (may be a subset of all rooms)
        actual_rooms = set(df['ห้องสอบ'].unique())
        
        # Filter and keep only rooms that have students, maintaining original order
        ordered_rooms = [room for room in program_room_list if room in actual_rooms]
        
        if not ordered_rooms:
            logger.error(f"No matching rooms found for program {program}")
            return False
            
        total_pages = len(ordered_rooms)
        academic_year = "2568"  # Thai year for 2024
        
        # Helper function to calculate the optimal font size for text to fit in width
        def get_optimal_font_size(text, width, min_size=4, max_size=14, pdf=pdf):
            if not text:
                return max_size
                
            # Start with the maximum size
            for font_size in range(max_size, min_size - 1, -1):
                pdf.set_font('thsarabun', '', font_size)
                text_width = pdf.get_string_width(str(text))
                if text_width <= width - 2:  # 2mm padding
                    return font_size
            
            # If we get here, even the smallest font size is too large, return min size
            return min_size
        
        # Process each room in the order from EXAM_ROOMS
        for room_idx, room in enumerate(ordered_rooms, 1):
            room_data = df[df['ห้องสอบ'] == room].copy()
            room_data = room_data.sort_values('เลขที่นั่งสอบ')
            
            # Add a new page for each room
            pdf.add_page()
            
            # Page number
            pdf.set_font('thsarabun', 'B', 10)
            pdf.cell(0, 4, f'แผ่นที่ {room_idx}/{total_pages}', align='R', ln=True)
            
            # Headers with fixed font sizes and heights
            pdf.set_font('thsarabun', 'B', 14)
            pdf.cell(0, 6, 'ใบรายชื่อผู้เข้าสอบ', align='C', ln=True)
            pdf.cell(0, 6, 'การสอบคัดเลือกเข้าศึกษาต่อชั้นมัธยมศึกษาปีที่ ' + ('4' if program.startswith('m4') else '1'), align='C', ln=True)
            
            # Handle long program names by wrapping text
            pdf.set_font('thsarabun', '', 12)
            program_name = PROGRAM_THAI_NAME[program]
            
            # Calculate available width (A4 width - margins)
            available_width = 190  # mm
            
            # Split text into lines that fit the width
            pdf.set_x(10)  # Reset X position
            lines = []
            current_line = ""
            words = program_name.split()
            
            for word in words:
                test_line = current_line + (" " if current_line else "") + word
                if pdf.get_string_width(test_line) <= available_width:
                    current_line = test_line
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = word
            
            if current_line:
                lines.append(current_line)
            
            # Output each line centered
            for line in lines:
                pdf.cell(0, 5, line, align='C', ln=True)
            
            pdf.set_font('thsarabun', '', 12)
            pdf.cell(0, 5, f'ปีการศึกษา {academic_year} โรงเรียนเบญจมราชูทิศ', align='C', ln=True)
            exam_date = "15 มีนาคม 2568" if program.startswith('m1') else "1 มีนาคม 2568"
            pdf.cell(0, 5, f'วันที่ {exam_date} ห้องสอบที่ {room_idx} ห้อง {room}', align='C', ln=True)
            
            # Define exact column widths - carefully tuned to fit portrait page
            # Ensure these add up to PDF page width minus margins
            # A4 width is 210mm minus left and right margin (~20mm total)
            headers = ['เลขที่นั่งสอบ', 'เลขประจำตัวสอบ', 'ชื่อ - นามสกุล', 'โรงเรียน', 'ลงชื่อผู้เข้าสอบ', 'หมายเหตุ']
            col_widths = [20, 20, 50, 50, 35, 20]  # Total = 195mm
            
            # Calculate start X position to center the table
            page_width = 210
            table_width = sum(col_widths)
            start_x = (page_width - table_width) / 2
            
            # Set start position with fixed spacing
            pdf.ln(2)
            pdf.set_x(start_x)
            
            # Draw table header with fixed height and font
            header_start_x = pdf.get_x()
            
            # Fixed header row height
            header_height = 6
            
            # Draw each header with appropriate font size
            current_x = header_start_x
            for i, (width, header) in enumerate(zip(col_widths, headers)):
                # Set position for this header cell
                pdf.set_xy(current_x, pdf.get_y())
                
                # Set different font sizes based on the header content
                if header in ['เลขที่นั่งสอบ', 'เลขประจำตัวสอบ']:
                    # Smaller font size for exam number headers
                    pdf.set_font('thsarabun', 'B', 10)
                else:
                    # Regular font size for other headers
                    pdf.set_font('thsarabun', 'B', 14)
                
                # Draw the header cell
                pdf.cell(width, header_height, header, border=1, align='C')
                
                # Move to next cell position
                current_x += width
            
            pdf.ln()
            
            # Table content with fixed row height - compact enough for 33 students
            row_height = 6.5  # Increased row height for better vertical padding
            
            for _, row in room_data.iterrows():
                # Reset to consistent starting X position for each row
                pdf.set_x(header_start_x)
                
                # Determine y position before drawing the row
                start_y = pdf.get_y()
                current_x = header_start_x
                
                # Prepare cell values
                cell_values = [
                    str(row['เลขที่นั่งสอบ']),
                    str(row['เลขประจำตัวสอบ']),
                    str(row['ชื่อ - นามสกุล']),
                    str(row['โรงเรียน']),
                    "",  # Empty signature column
                    ""   # Empty note column
                ]
                
                # Draw each cell in the row
                for i, (width, value) in enumerate(zip(col_widths, cell_values)):
                    # Set position to current cell
                    pdf.set_xy(current_x, start_y)
                    
                    # Cell borders
                    pdf.cell(width, row_height, "", border=1)
                    
                    # Skip empty cells
                    if not value:
                        current_x += width
                        continue
                    
                    # Determine the optimal font size for this content
                    align = 'C' if i in [0, 1, 4, 5] else 'L'
                    padding = 1 if align == 'L' else 0
                    
                    # Set an appropriate font size that makes the text fit
                    font_size = get_optimal_font_size(value, width - (2 * padding))
                    pdf.set_font('thsarabun', '', font_size)
                    
                    # Add text content - precisely centered vertically with more padding
                    # Keep content_height the same but increase the overall row height
                    content_height = 3.5
                    y_position = start_y + (row_height - content_height) / 2
                    pdf.set_xy(current_x + padding, y_position)
                    pdf.cell(width - (2 * padding), content_height, value, align=align)
                    
                    # Move to next cell position
                    current_x += width
                
                # Move to next row - exactly at the end of this row
                pdf.set_y(start_y + row_height)
        
        # Save the PDF
        pdf.output(output_file)
        logger.info(f"Successfully created PDF for program {program}: {output_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error creating PDF for program {program}: {str(e)}")
        return False

def main():
    """Main function to process student exam room assignments"""
    logger.info("Starting exam room assignment process")
    output_excel = "exam_room_assignments.xlsx"
    
    try:
        # Remove existing output file if it exists
        if Path(output_excel).exists():
            Path(output_excel).unlink()
            logger.info(f"Removed existing output file: {output_excel}")
        
        # Load and process data
        data = load_and_prepare_data(EXCEL_INPUT_FILE, SHEET_NAME)
        if data is None:
            logger.error("Failed to load input data")
            return
        
        # Process each program
        for program in EXAM_ID_PREFIX.keys():
            try:
                program_data = data[data["program"] == program]
                if len(program_data) == 0:
                    logger.info(f"No students found for program {program}")
                    continue
                
                # Process the data
                program_data = clean_school_names(program_data)
                program_data = format_student_names(program_data)
                program_data = assign_exam_details(program_data)
                
                if program_data is None:
                    logger.error(f"Failed to process program {program}")
                    continue
                
                # Prepare final dataframe and save
                final_df = prepare_final_dataframe(program_data)
                
                if save_to_excel(final_df, program, output_excel):
                    logger.info(f"Successfully processed {len(final_df)} students for program {program}")
                    
                    # Generate PDF with a unique filename for each program
                    pdf_filename = f"exam_room_{program}.pdf"
                    if create_pdf_file(final_df, program, pdf_filename):
                        logger.info(f"Successfully created PDF for program {program}: {pdf_filename}")
                
            except Exception as e:
                logger.error(f"Error processing program {program}: {str(e)}")
                continue
        
        logger.info(f"Exam room assignment process completed. Results saved to {output_excel} and PDF files")
        
    except Exception as e:
        logger.error(f"Critical error in main process: {str(e)}")

if __name__ == "__main__":
    main()




